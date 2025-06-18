# import openpyxl
# from openpyxl.worksheet.table import Table, TableStyleInfo
# import os

# # נתיב לקובץ
# excel_path = r"C:\Users\idanp\OneDrive\שולחן העבודה\automation\practice.xlsx"

# # אם הקובץ קיים – טוען אותו, אחרת יוצר חדש
# if os.path.exists(excel_path):
#     wb = openpyxl.load_workbook(excel_path)
#     sheet = wb.active
# else:
#     wb = openpyxl.Workbook()
#     sheet = wb.active

# # כתיבת כותרות לעמודות A ו־B
# sheet['A1'] = "name"
# sheet['B1'] = "last name"

# # יצירת טבלה על טווח A1:B2 (גם אם אין שורות נוספות, חובה להגדיר לפחות שורה אחת)
# table = Table(displayName="practice", ref="A1:B2")

# # עיצוב לטבלה
# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
# table.tableStyleInfo = style

# # הוספת הטבלה לגיליון
# sheet.add_table(table)

# # שמירת הקובץ
# wb.save(excel_path)
# print("✔️ הקובץ עודכן ונשמר בהצלחה.")

# ##### הוספת שמות שחקני "חברים" והגדלת התא פלוס יישור למרכז
# import openpyxl
# from openpyxl.worksheet.table import Table, TableStyleInfo
# from openpyxl.styles import Alignment
# import os

# # נתיב לקובץ האקסל
# excel_path = r"C:\Users\idanp\OneDrive\שולחן העבודה\automation\practice.xlsx"

# # פתיחת קובץ אם קיים, או יצירת חדש
# if os.path.exists(excel_path):
#     wb = openpyxl.load_workbook(excel_path)
#     sheet = wb.active
# else:
#     wb = openpyxl.Workbook()
#     sheet = wb.active

# # שמות הדמויות מהסדרה "חברים"
# friends_names = [
#     ("Rachel", "Green"),
#     ("Monica", "Geller"),
#     ("Phoebe", "Buffay"),
#     ("Joey", "Tribbiani"),
#     ("Chandler", "Bing"),
#     ("Ross", "Geller")
# ]

# # כתיבת כותרות
# sheet['A1'] = "name"
# sheet['B1'] = "last name"

# # הכנסת השמות
# for i, (first, last) in enumerate(friends_names, start=2):
#     sheet[f"A{i}"] = first
#     sheet[f"B{i}"] = last

# # יישור למרכז + התאמת גובה ורוחב
# alignment = Alignment(horizontal='center', vertical='center')
# for row in sheet.iter_rows(min_row=1, max_row=len(friends_names) + 1, min_col=1, max_col=2):
#     for cell in row:
#         cell.alignment = alignment

# # קביעת גובה שורה 25
# for i in range(1, len(friends_names) + 2):
#     sheet.row_dimensions[i].height = 25

# # קביעת רוחב עמודות
# sheet.column_dimensions['A'].width = 18
# sheet.column_dimensions['B'].width = 18

# # יצירת טבלה
# table_range = f"A1:B{len(friends_names)+1}"
# table = Table(displayName="practice", ref=table_range)

# # עיצוב טבלה
# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
# table.tableStyleInfo = style

# # מחיקת טבלאות קיימות אם יש
# if sheet.tables:
#     for tbl in list(sheet.tables):
#         del sheet.tables[tbl]

# sheet.add_table(table)

# # שמירת הקובץ
# wb.save(excel_path)
# print("✔️ הטבלה נשמרה עם עיצוב ומידע.")

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

# הנתיב לקובץ אקסל
excel_path = r"C:\Users\idanp\OneDrive\שולחן העבודה\automation\practice.xlsx"

# טוענים את הקובץ
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

# קובעים את גבולות הטבלה
min_col = sheet.min_column
max_col = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

# יוצרים טבלה
table = Table(displayName="practice", ref=f"{sheet.cell(min_row, min_col).coordinate}:{sheet.cell(max_row, max_col).coordinate}")

# עיצוב טבלה
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style

# מוסיפים את הטבלה לגיליון
sheet.add_table(table)

# שמירה
wb.save(excel_path)
print("✅ ה.")
